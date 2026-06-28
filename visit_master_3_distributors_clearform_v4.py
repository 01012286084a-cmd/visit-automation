import os
import re
from datetime import datetime, date, timedelta
from openpyxl import load_workbook
from playwright.sync_api import sync_playwright

# ============================================================
# CONFIG
# ============================================================

FORM_URL = "https://forms.office.com/Pages/ResponsePage.aspx?id=kA3onZv_-UK1DjonUZ3sfA2XrmPVXLJIlfHya3C-kcBUNVY1WkxNWjFMVTFEMDVSSU5FVkJPR0MyQS4u&origin=Invitation&channel=0"

BASE_DIR = r"D:\Automate\FormVisits" if not CLOUD_MODE else os.path.join(os.path.dirname(__file__), "data")
USER_DATA_DIR = os.path.join(BASE_DIR, "user_data")
SCREENSHOT_DIR = os.path.join(BASE_DIR, "screenshots")
LOG_DIR = os.path.join(BASE_DIR, "logs")

EXCEL_FILE = r"D:\OneDrive - The Savola Group\visits\VISIT DIST.xlsm" if not CLOUD_MODE else os.path.join(BASE_DIR, "VISIT DIST.xlsm")
SCHEDULE_SHEET_NAME = "Schedule"
SCHEDULE_CODE_COLUMN = 3  # Column C

# ملفات مشتركة بين كل الزيارات
_DATA_DIR = r"D:\OneDrive - The Savola Group\AuditDistCloud\Input" if not CLOUD_MODE else BASE_DIR
MARKET_CREDIT_FILE = os.path.join(_DATA_DIR, "Market credit 2022 v2 Delta A.xlsx")
EXPENSES_APPROVAL_FILE = os.path.join(_DATA_DIR, "Expenses Aproval.xlsx")
CR_NOTE_FILE = os.path.join(_DATA_DIR, "CR. Note.xlsx")


# ============================================================
# SAFETY SWITCHES
# ============================================================

# خليه False طول مرحلة الاختبار.
# في السحابة بتتظبط auto من GitHub Actions (env: SUBMIT=true)
SUBMIT = os.environ.get("SUBMIT", "false").lower() == "true"

# تعليم مربع إرسال نسخة على الإيميل
SEND_EMAIL_RECEIPT = True

# استخدام Clear Form الموجود في أعلى الفورم قبل التعبئة
CLEAR_FORM_BEFORE_FILL = True

# لو Clear Form فشل أو المرفقات القديمة فضلت ظاهرة: يوقف فورًا لمنع التكرار
STRICT_CLEAR_FORM_REQUIRED = True

# مع STRICT_CLEAR_FORM_REQUIRED=True لا ننصح بالـ fallback لأنه لا يمسح المرفقات
FALLBACK_MANUAL_CLEAR_IF_TOP_CLEAR_FAILS = False

# مود السحابة: بيخفي الـ input() وبيستعمل headless browser
# GitHub Actions بيظبط automatic، لو شغال محليًا غيّره لـ False
CLOUD_MODE = os.environ.get("GITHUB_ACTIONS") == "true"

# روابط تحميل الملفات من OneDrive للسحابة (لو الملف مش موجود محلياً)
# اعمل share link لكل ملف وضيفه هنا — المفتاح هو المسار المحلي، القيمة هي رابط التحميل
REMOTE_FILES = {}

VISIT_DATE = f"{datetime.now().month}/{datetime.now().day}/{datetime.now().year}"
RUN_TIME = datetime.now().strftime("%Y%m%d_%H%M%S")

os.makedirs(SCREENSHOT_DIR, exist_ok=True)
os.makedirs(LOG_DIR, exist_ok=True)

CURRENT_CODE = "UNKNOWN"
SCREENSHOT_FILE = os.path.join(SCREENSHOT_DIR, f"visit_master_{CURRENT_CODE}_{RUN_TIME}.png")
LOG_FILE = os.path.join(LOG_DIR, f"visit_master_{CURRENT_CODE}_{RUN_TIME}.log")

# ============================================================
# VISIT PROFILES - BLQ / AGA / DAM
# ============================================================

_VISITS_DIR = r"D:\OneDrive - The Savola Group\visits" if not CLOUD_MODE else BASE_DIR

VISIT_PROFILES = {
    "BLQ": {
        "sheet_name": "BLQ",
        "distributor": "Dakahlia - Al Yasmine",
        "mk_file": os.path.join(_VISITS_DIR, "MKBLQ.jpeg"),
        "kh_file": os.path.join(_VISITS_DIR, "KHBLQ.jpeg"),
    },
    "AGA": {
        "sheet_name": "AGA",
        "distributor": "Dakahlia Aga – Al Yasmine",
        "mk_file": os.path.join(_VISITS_DIR, "MKAGA.jpeg"),
        "kh_file": os.path.join(_VISITS_DIR, "KHAGA.jpeg"),
    },
    "DAM": {
        "sheet_name": "DAM",
        "distributor": "Damietta - Khaled Ismaiel Al-Fath",
        "mk_file": os.path.join(_VISITS_DIR, "MKDAM.jpeg"),
        "kh_file": os.path.join(_VISITS_DIR, "KHDAM.jpeg"),
    },
}

# ============================================================
# REMOTE FILES REGISTRY
# ============================================================

# الملفات دلوقتي جوه data/ في نفس الـ repo — مش محتاج URLs
REMOTE_FILES[EXCEL_FILE] = ""
REMOTE_FILES[MARKET_CREDIT_FILE] = ""
REMOTE_FILES[EXPENSES_APPROVAL_FILE] = ""
REMOTE_FILES[CR_NOTE_FILE] = ""
for _code, _profile in VISIT_PROFILES.items():
    REMOTE_FILES[_profile["mk_file"]] = ""
    REMOTE_FILES[_profile["kh_file"]] = ""


def ensure_file_available(file_path):
    """لو الملف مش موجود محلياً، نحمله من رابط OneDrive"""
    if os.path.exists(file_path):
        return file_path

    url = REMOTE_FILES.get(file_path, "")
    if not url:
        raise FileNotFoundError(
            f"File not found locally and no remote URL configured: {file_path}\n"
            f"Add its OneDrive share link to REMOTE_FILES in the script."
        )

    log(f"Downloading {file_path} from {url}")
    os.makedirs(os.path.dirname(file_path), exist_ok=True)

    try:
        import requests
        resp = requests.get(url, timeout=120, allow_redirects=True)
        resp.raise_for_status()
        with open(file_path, "wb") as f:
            f.write(resp.content)
        log(f"Downloaded successfully: {file_path} ({len(resp.content)} bytes)")
        return file_path
    except ImportError:
        raise Exception("requests library required for downloading files. Run: pip install requests")
    except Exception as e:
        raise Exception(f"Failed to download {file_path} from {url}: {e}")


# ============================================================
# LOGGING + HELPERS
# ============================================================

def set_run_files(code):
    global CURRENT_CODE, SCREENSHOT_FILE, LOG_FILE
    CURRENT_CODE = code
    SCREENSHOT_FILE = os.path.join(SCREENSHOT_DIR, f"visit_master_{code}_{RUN_TIME}.png")
    LOG_FILE = os.path.join(LOG_DIR, f"visit_master_{code}_{RUN_TIME}.log")


def log(msg):
    line = f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] {msg}"
    print(line)
    with open(LOG_FILE, "a", encoding="utf-8") as f:
        f.write(line + "\n")


def normalize_text(s):
    if s is None:
        return ""
    s = str(s).replace("\u00a0", " ").strip()
    s = re.sub(r"\s+", " ", s)
    return s


def normalize_code(value):
    code = normalize_text(value)
    code = code.strip('"').strip("'")
    if code.lower().endswith(".py"):
        code = code[:-3]
    return code.upper().strip()


def cell_to_text(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y")
    return str(value).strip()


def read_range_text(ws, range_address):
    values = []
    for row in ws[range_address]:
        for cell in row:
            txt = cell_to_text(cell.value)
            if txt:
                values.append(txt)
    return "\n".join(values).strip()


def get_value_from_mapping(ws, cfg):
    if "fixed" in cfg:
        return str(cfg.get("fixed", "")).strip()
    if "source_range" in cfg:
        return read_range_text(ws, cfg["source_range"])
    return ""


def excel_value_to_date(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        try:
            base = datetime(1899, 12, 30)
            return (base + timedelta(days=int(value))).date()
        except Exception:
            return None

    text = normalize_text(value)
    if not text:
        return None

    # لو النص فيه وقت بعد التاريخ
    text = text.split(" ")[0]

    possible_formats = [
        "%d/%m/%Y",
        "%m/%d/%Y",
        "%Y-%m-%d",
        "%d-%m-%Y",
        "%m-%d-%Y",
        "%d.%m.%Y",
        "%m.%d.%Y",
    ]

    for fmt in possible_formats:
        try:
            return datetime.strptime(text, fmt).date()
        except Exception:
            pass
    return None

# ============================================================
# SCHEDULE READER (openpyxl — no Excel required)
# ============================================================

def refresh_workbook_and_get_today_code():
    """
    يقرأ شيت Schedule من ملف Excel مباشرة بـ openpyxl
    (بدون COM — شغال على Linux)
    ملاحظة: Power Query / Pivot Refresh يحتاج Excel نفسه؛
    تأكد أن الملف محدث قبل رفعه للسحابة.
    """
    ensure_file_available(EXCEL_FILE)

    today = datetime.now().date()
    log(f"Today date = {today.strftime('%d/%m/%Y')}")

    wb = load_workbook(EXCEL_FILE, data_only=True, keep_vba=True)

    if SCHEDULE_SHEET_NAME not in wb.sheetnames:
        raise Exception(f"Sheet not found: {SCHEDULE_SHEET_NAME}. Available: {wb.sheetnames}")

    ws = wb[SCHEDULE_SHEET_NAME]
    matches = []

    for row in range(1, ws.max_row + 1):
        row_has_today = False
        for col in range(1, ws.max_column + 1):
            cell_date = excel_value_to_date(ws.cell(row, col).value)
            if cell_date == today:
                row_has_today = True
                break

        if row_has_today:
            raw_code = ws.cell(row, SCHEDULE_CODE_COLUMN).value
            code = normalize_code(raw_code)
            log(f"Today's date found in row {row}. Column C raw = {raw_code}, code = {code}")
            if code:
                matches.append((row, code))
            else:
                log(f"WARNING: Row {row} has today's date but column C is empty.")

    if not matches:
        raise Exception(f"No visit code found for today's date {today.strftime('%d/%m/%Y')} in Schedule.")

    if len(matches) > 1:
        log("WARNING: More than one row found for today. First will be used.")
        for row, code in matches:
            log(f"Found: row={row}, code={code}")

    selected_code = matches[0][1]
    log(f"Selected visit code: {selected_code}")

    if selected_code not in VISIT_PROFILES:
        valid_codes = ", ".join(sorted(VISIT_PROFILES.keys()))
        raise Exception(f"Invalid visit code in Schedule column C: {selected_code}. Valid codes: {valid_codes}")

    return selected_code

# ============================================================
# FORM MAP BUILDER
# ============================================================

def build_form_map(profile):
    return {
        1:  {"type": "radio",    "fixed": profile["distributor"]},
        2:  {"type": "radio",    "fixed": "أحمد السمنودى"},
        3:  {"type": "date",     "fixed": VISIT_DATE},
        4:  {"type": "textarea", "source_range": "D4:D20"},
        5:  {"type": "radio",    "fixed": "الارصدة مطابقة"},
        6:  {"type": "radio",    "fixed": "الارصدة مطابقة"},
        7:  {"type": "file",     "file": ""},
        8:  {"type": "file",     "file": profile["mk_file"]},
        9:  {"type": "textarea", "source_range": "D23:D33"},
        10: {"type": "radio",    "fixed": "الارصدة مطابقة"},
        11: {"type": "file",     "file": ""},
        12: {"type": "file",     "file": profile["kh_file"]},
        13: {"type": "textarea", "source_range": "D36:D68"},
        14: {"type": "radio",    "fixed": "الارصدة مطابقة"},
        15: {"type": "file",     "file": MARKET_CREDIT_FILE},
        16: {"type": "file",     "file": ""},
        17: {"type": "textarea", "source_range": "D71:D82"},
        18: {"type": "radio",    "fixed": "مطابق"},
        19: {"type": "file",     "file": EXPENSES_APPROVAL_FILE},
        20: {"type": "file",     "file": ""},
        21: {"type": "textarea", "source_range": "D85:D106"},
        22: {"type": "file",     "file": CR_NOTE_FILE},
        23: {"type": "file",     "file": ""},
        24: {"type": "radio",    "fixed": "مطابق"},
        25: {"type": "textarea", "source_range": "D109:D127"},
        26: {"type": "file",     "file": ""},
        27: {"type": "textarea", "source_range": "D130:D143"},
        28: {"type": "radio",    "fixed": "قائمة بيضاء"},
        29: {"type": "radio",    "fixed": "7"},
        30: {"type": "textarea", "source_range": "D145:D167"},
        31: {"type": "radio",    "fixed": "10"},
        32: {"type": "textarea", "source_range": "D169:D188"},
        33: {"type": "radio",    "fixed": "6"},
        34: {"type": "textarea", "source_range": "D190:D213"},
        35: {"type": "radio",    "fixed": "10"},
        36: {"type": "textarea", "source_range": "D215:D231"},
        37: {"type": "radio",    "fixed": "6"},
        38: {"type": "textarea", "source_range": "D233:D250"},
        39: {"type": "radio",    "fixed": "10"},
        40: {"type": "textarea", "source_range": "D252:D264"},
        41: {"type": "radio",    "fixed": "7"},
        42: {"type": "textarea", "source_range": "D266:D278"},
        43: {"type": "textarea", "fixed": "ملتزم"},
        44: {"type": "radio",    "fixed": "7"},
    }

# ============================================================
# PLAYWRIGHT FORM HELPERS
# ============================================================

def wait_until_questions_ready(page):
    log("Waiting until form questions are ready...")

    for attempt in range(1, 61):
        questions = page.locator('[data-automation-id="questionItem"]')
        count = questions.count()

        if count > 0:
            log(f"Questions are ready. Total questions found: {count}")
            return count

        log(f"Questions not ready yet. Attempt {attempt}/60")

        start_locators = [
            page.get_by_role("button", name=re.compile(r"Start now|START NOW|Start|ابدأ|بدء", re.I)),
            page.locator('button:has-text("Start now")'),
            page.locator('button:has-text("START NOW")'),
            page.locator('text=Start now'),
            page.locator('text=START NOW'),
        ]

        for btn in start_locators:
            try:
                if btn.count() > 0 and btn.first.is_visible():
                    log("START NOW page detected. Clicking START NOW...")
                    btn.first.click(timeout=10000)
                    page.wait_for_timeout(5000)
                    break
            except Exception:
                pass

        page.wait_for_timeout(1000)

    fail_shot = os.path.join(SCREENSHOT_DIR, f"questions_not_found_{CURRENT_CODE}_{RUN_TIME}.png")

    try:
        page.screenshot(path=fail_shot, full_page=True)
    except Exception:
        pass

    try:
        body_text = page.locator("body").inner_text(timeout=5000)
    except Exception:
        body_text = ""

    log(f"ERROR: Questions not found after waiting. Screenshot: {fail_shot}")
    log(f"Body preview: {body_text[:1500]}")

    raise Exception("Form questions did not load. START NOW may not have been clicked or page did not continue.")


def get_question(page, question_number):
    questions = page.locator('[data-automation-id="questionItem"]')
    count = questions.count()

    if question_number < 1 or question_number > count:
        raise Exception(f"Question {question_number} not found. Total questions found: {count}")

    return questions.nth(question_number - 1)


def scroll_to_question(page, question_number):
    question = get_question(page, question_number)
    question.scroll_into_view_if_needed(timeout=10000)
    page.wait_for_timeout(700)
    return question

# ============================================================
# CLEAR FORM FIRST - REAL SELECTORS FROM DIAGNOSTIC
# ============================================================

def get_old_file_markers():
    """
    أسماء الملفات التي لو ظهرت بعد Clear Form معناها المسح لم ينجح.
    """
    return [
        "MKBLQ.jpeg",
        "MKBLQ 2.jpeg",
        "MKAGA.jpeg",
        "MKAGA 2.jpeg",
        "MKDAM.jpeg",
        "MKDAM 2.jpeg",
        "KHBLQ.jpeg",
        "KHBLQ 2.jpeg",
        "KHAGA.jpeg",
        "KHAGA 2.jpeg",
        "KHDAM.jpeg",
        "KHDAM 2.jpeg",
        "Market credit 2022 v2 Delta A.xlsx",
        "Market credit 2022 v2 Delta A 2.xlsx",
        "Expenses Aproval.xlsx",
        "Expenses Aproval 2.xlsx",
        "CR. Note.xlsx",
        "CR. Note 2.xlsx",
    ]


def collect_visible_text_by_scrolling(page):
    """
    يجمع النص الظاهر بعد النزول على أجزاء الصفحة، للتأكد من عدم ظهور مرفقات قديمة.
    """
    collected = []

    try:
        page.evaluate("window.scrollTo(0, 0)")
        page.wait_for_timeout(800)
    except Exception:
        pass

    for _ in range(1, 28):
        try:
            body_text = page.locator("body").inner_text(timeout=5000)
            if body_text:
                collected.append(body_text)
        except Exception:
            pass

        try:
            page.mouse.wheel(0, 900)
            page.wait_for_timeout(350)
        except Exception:
            pass

    try:
        page.evaluate("window.scrollTo(0, 0)")
        page.wait_for_timeout(800)
    except Exception:
        pass

    return "\n".join(collected)


def click_top_clear_form(page):
    """
    يضغط زر Form Menu الحقيقي، ثم Clear Form، ثم زر تأكيد Clear Form.
    مبني على التشخيص الحقيقي:
    - button[aria-label="Form Menu"]
    - role=menuitem text=Clear Form
    - button[aria-label="Clear Form"] داخل Dialog التأكيد
    """
    log("Trying to clear form using REAL selectors: Form Menu > Clear Form > Confirm Clear Form")

    try:
        page.evaluate("window.scrollTo(0, 0)")
        page.wait_for_timeout(1000)
    except Exception:
        pass

    # 1) اضغط زر الثلاث نقاط الحقيقي
    menu_button = page.locator('button[aria-label="Form Menu"]')

    if menu_button.count() == 0:
        fail_shot = os.path.join(SCREENSHOT_DIR, f"form_menu_button_not_found_{CURRENT_CODE}_{RUN_TIME}.png")
        try:
            page.screenshot(path=fail_shot, full_page=True)
        except Exception:
            pass
        log(f"ERROR: Form Menu button not found. Screenshot: {fail_shot}")
        return False

    try:
        menu_button.first.scroll_into_view_if_needed(timeout=10000)
        page.wait_for_timeout(300)
        menu_button.first.click(force=True, timeout=10000)
        page.wait_for_timeout(1200)
        log("Form Menu button clicked successfully")
    except Exception as e:
        fail_shot = os.path.join(SCREENSHOT_DIR, f"form_menu_button_click_failed_{CURRENT_CODE}_{RUN_TIME}.png")
        try:
            page.screenshot(path=fail_shot, full_page=True)
        except Exception:
            pass
        log(f"ERROR: Failed to click Form Menu button: {e}. Screenshot: {fail_shot}")
        return False

    # 2) اضغط Clear Form من القائمة
    clear_menu_item = page.get_by_role("menuitem", name=re.compile(r"^Clear Form$", re.I))

    if clear_menu_item.count() == 0:
        clear_menu_item = page.locator('[role="menuitem"]').filter(has_text=re.compile(r"^Clear Form$", re.I))

    if clear_menu_item.count() == 0:
        fail_shot = os.path.join(SCREENSHOT_DIR, f"clear_form_menuitem_not_found_{CURRENT_CODE}_{RUN_TIME}.png")
        try:
            page.screenshot(path=fail_shot, full_page=True)
        except Exception:
            pass
        log(f"ERROR: Clear Form menuitem not found after opening Form Menu. Screenshot: {fail_shot}")
        return False

    try:
        clear_menu_item.first.click(force=True, timeout=10000)
        page.wait_for_timeout(1500)
        log("Clear Form menuitem clicked successfully")
    except Exception as e:
        fail_shot = os.path.join(SCREENSHOT_DIR, f"clear_form_menuitem_click_failed_{CURRENT_CODE}_{RUN_TIME}.png")
        try:
            page.screenshot(path=fail_shot, full_page=True)
        except Exception:
            pass
        log(f"ERROR: Failed to click Clear Form menuitem: {e}. Screenshot: {fail_shot}")
        return False

    # 3) اضغط زر التأكيد النهائي Clear Form
    page.wait_for_timeout(1500)

    confirm_button = page.locator('button[aria-label="Clear Form"]')

    if confirm_button.count() == 0:
        confirm_button = page.get_by_role("button", name=re.compile(r"^Clear Form$", re.I))

    if confirm_button.count() == 0:
        fail_shot = os.path.join(SCREENSHOT_DIR, f"clear_form_confirm_button_not_found_{CURRENT_CODE}_{RUN_TIME}.png")
        try:
            page.screenshot(path=fail_shot, full_page=True)
        except Exception:
            pass
        log(f"ERROR: Confirm Clear Form button not found. Screenshot: {fail_shot}")
        return False

    try:
        confirm_button.last.click(force=True, timeout=10000)
        page.wait_for_timeout(4000)
        log("Confirm Clear Form button clicked successfully")
    except Exception as e:
        fail_shot = os.path.join(SCREENSHOT_DIR, f"clear_form_confirm_click_failed_{CURRENT_CODE}_{RUN_TIME}.png")
        try:
            page.screenshot(path=fail_shot, full_page=True)
        except Exception:
            pass
        log(f"ERROR: Failed to click Confirm Clear Form button: {e}. Screenshot: {fail_shot}")
        return False

    # 4) انتظر الفورم يرجع بعد المسح
    try:
        page.wait_for_timeout(3000)
        wait_until_questions_ready(page)
    except Exception as e:
        log(f"WARNING: Questions readiness check after Clear Form had issue: {e}")

    # 5) تحقق أن المرفقات القديمة اختفت بالـ scrolling
    full_text_after_clear = collect_visible_text_by_scrolling(page)
    still_found = []

    for marker in get_old_file_markers():
        if marker in full_text_after_clear:
            still_found.append(marker)

    if still_found:
        fail_shot = os.path.join(SCREENSHOT_DIR, f"clear_form_failed_files_still_visible_{CURRENT_CODE}_{RUN_TIME}.png")
        try:
            page.screenshot(path=fail_shot, full_page=True)
        except Exception:
            pass
        log(f"WARNING: Clear Form confirmed but old files still visible: {still_found}")
        log(f"Screenshot: {fail_shot}")
        return False

    try:
        page.evaluate("window.scrollTo(0, 0)")
        page.wait_for_timeout(800)
    except Exception:
        pass

    log("Clear Form completed successfully using real selectors and no old files are visible")
    return True


def clear_all_text_fields(page):
    log("Fallback: Clearing text/date fields...")

    fields = page.locator(
        'textarea, '
        'input[type="text"], '
        'input[data-automation-id="textInput"], '
        'input[placeholder="Enter your answer"], '
        'input[aria-label="Date picker"]'
    )

    count = fields.count()
    cleared = 0

    for i in range(count):
        field = fields.nth(i)
        try:
            if not field.is_visible():
                continue
            field.scroll_into_view_if_needed(timeout=5000)
            page.wait_for_timeout(100)
            field.click(timeout=5000)
            page.keyboard.press("Control+A")
            page.keyboard.press("Backspace")
            page.wait_for_timeout(100)
            cleared += 1
        except Exception:
            pass

    log(f"Fallback text/date fields cleared: {cleared}")


def clear_all_radio_selections(page):
    log("Fallback: Clearing radio selections using Clear Selection buttons...")

    total_clicked = 0

    for round_no in range(1, 6):
        clicked_this_round = 0

        clear_buttons = page.locator(
            'button[aria-label="Clear Selection"], '
            '[aria-label="Clear Selection"], '
            'button:has-text("Clear Selection")'
        )

        count = clear_buttons.count()

        for i in range(count):
            btn = clear_buttons.nth(i)
            try:
                if not btn.is_visible():
                    continue
                btn.scroll_into_view_if_needed(timeout=5000)
                page.wait_for_timeout(100)
                btn.click(force=True, timeout=5000)
                page.wait_for_timeout(300)
                clicked_this_round += 1
                total_clicked += 1
            except Exception:
                pass

        if clicked_this_round == 0:
            break

        log(f"Fallback clear selection round {round_no}: clicked {clicked_this_round}")

    log(f"Fallback total Clear Selection clicked: {total_clicked}")


def clear_all_checkboxes(page):
    log("Fallback: Clearing checkboxes...")

    checkboxes = page.locator('input[type="checkbox"], [role="checkbox"]')
    count = checkboxes.count()
    unchecked = 0

    for i in range(count):
        cb = checkboxes.nth(i)
        try:
            if not cb.is_visible():
                continue

            checked = False
            try:
                checked = cb.is_checked()
            except Exception:
                aria_checked = normalize_text(cb.get_attribute("aria-checked"))
                checked = aria_checked.lower() == "true"

            if checked:
                try:
                    cb.uncheck(force=True, timeout=5000)
                except Exception:
                    cb.click(force=True, timeout=5000)
                page.wait_for_timeout(200)
                unchecked += 1
        except Exception:
            pass

    log(f"Fallback checkboxes unchecked: {unchecked}")


def clear_form_before_fill(page):
    if not CLEAR_FORM_BEFORE_FILL:
        log("CLEAR_FORM_BEFORE_FILL = False -> skipping clear step")
        return

    log("========== START CLEAR FORM BEFORE FILL ==========")

    top_clear_ok = False

    try:
        top_clear_ok = click_top_clear_form(page)
    except Exception as e:
        log(f"WARNING: Error while clicking top Clear Form: {e}")
        top_clear_ok = False

    if top_clear_ok:
        page.wait_for_timeout(1500)
        wait_until_questions_ready(page)
        log("Top Clear Form completed successfully")
    else:
        log("Top Clear Form did not complete successfully")

        # تحقق: هل فيه داتا قديمة أصلاً ولا الفورم فاضي؟
        old_data_found = False
        try:
            full_text = collect_visible_text_by_scrolling(page)
            old_data_found = any(marker in full_text for marker in get_old_file_markers())
        except Exception:
            pass

        if not old_data_found:
            log("No old data found in form - nothing to clear, continuing normally")
        else:
            if STRICT_CLEAR_FORM_REQUIRED:
                fail_shot = os.path.join(SCREENSHOT_DIR, f"strict_clear_form_failed_{CURRENT_CODE}_{RUN_TIME}.png")
                try:
                    page.screenshot(path=fail_shot, full_page=True)
                except Exception:
                    pass
                log(f"ERROR: STRICT_CLEAR_FORM_REQUIRED=True, stopping to avoid duplicate attachments. Screenshot: {fail_shot}")
                raise Exception("Clear Form failed or old attachments are still visible. Script stopped to avoid duplicate attachments.")

            if FALLBACK_MANUAL_CLEAR_IF_TOP_CLEAR_FAILS:
                log("Fallback manual clear is enabled")
                try:
                    clear_all_text_fields(page)
                except Exception as e:
                    log(f"Warning while fallback clearing text fields: {e}")
                try:
                    clear_all_radio_selections(page)
                except Exception as e:
                    log(f"Warning while fallback clearing radio selections: {e}")
                try:
                    clear_all_checkboxes(page)
                except Exception as e:
                    log(f"Warning while fallback clearing checkboxes: {e}")
            else:
                raise Exception("Top Clear Form failed and fallback manual clear is disabled.")

    page.wait_for_timeout(1000)
    log("========== END CLEAR FORM BEFORE FILL ==========")

# ============================================================
# FILL QUESTION TYPES
# ============================================================

def fill_textarea_question(page, question_number, text_value):
    text_value = str(text_value or "").strip()

    if not text_value:
        log(f"Q{question_number}: text value is empty -> skipped")
        return

    question = scroll_to_question(page, question_number)

    text_locator = question.locator('textarea[data-automation-id="textInput"], textarea')

    if text_locator.count() == 0:
        text_locator = question.locator(
            'input[data-automation-id="textInput"], '
            'input[aria-label="Single Line Text"], '
            'input[placeholder="Enter your answer"], '
            'input[type="text"]:not([aria-label="Date picker"])'
        )

    if text_locator.count() == 0:
        text_locator = question.locator(
            'input:not([type="radio"]):not([type="file"]):not([aria-label="Date picker"])'
        )

    if text_locator.count() == 0:
        try:
            q_text = question.inner_text(timeout=3000)
        except Exception:
            q_text = ""

        fail_shot = os.path.join(SCREENSHOT_DIR, f"q{question_number}_text_input_not_found_{CURRENT_CODE}_{RUN_TIME}.png")
        try:
            page.screenshot(path=fail_shot, full_page=True)
        except Exception:
            pass
        log(f"Q{question_number}: Text input not found. Screenshot: {fail_shot}")
        log(f"Q{question_number}: Question text preview: {q_text[:1000]}")
        raise Exception(f"Q{question_number}: Text input not found")

    text_input = text_locator.first
    text_input.scroll_into_view_if_needed(timeout=10000)
    page.wait_for_timeout(300)

    text_input.click(timeout=10000)
    page.keyboard.press("Control+A")
    page.keyboard.press("Backspace")
    text_input.fill(text_value, timeout=10000)
    page.wait_for_timeout(300)

    log(f"Q{question_number}: text field filled ({len(text_value)} chars)")


def fill_date_question(page, question_number, date_value):
    date_value = str(date_value or "").strip()

    if not date_value:
        log(f"Q{question_number}: date value is empty -> skipped")
        return

    question = scroll_to_question(page, question_number)

    date_locator = question.locator(
        'input[aria-label="Date picker"], '
        'input[placeholder*="date"], '
        'input[type="text"]'
    )

    if date_locator.count() == 0:
        raise Exception(f"Q{question_number}: Date input not found")

    date_input = date_locator.first
    date_input.click(timeout=10000)
    page.keyboard.press("Control+A")
    page.keyboard.press("Backspace")
    date_input.fill(date_value, timeout=10000)
    page.keyboard.press("Tab")
    page.wait_for_timeout(300)

    log(f"Q{question_number}: date filled = {date_value}")


def select_radio_in_question(page, question_number, option_value):
    option_value = normalize_text(option_value)

    if not option_value:
        log(f"Q{question_number}: radio option is empty -> skipped")
        return

    question = scroll_to_question(page, question_number)

    radios = question.locator('input[type="radio"]')
    radio_count = radios.count()

    for i in range(radio_count):
        radio = radios.nth(i)
        val = normalize_text(radio.get_attribute("value"))

        if val == option_value:
            radio.check(force=True, timeout=10000)
            page.wait_for_timeout(400)
            log(f"Q{question_number}: radio selected by input value = {option_value}")
            return

    choices = question.locator('[data-automation-id="choiceItem"]')

    for i in range(choices.count()):
        choice = choices.nth(i)
        try:
            txt = normalize_text(choice.inner_text(timeout=1000))
        except Exception:
            txt = ""

        if txt == option_value:
            choice.click(timeout=10000)
            page.wait_for_timeout(400)
            log(f"Q{question_number}: radio selected by choice text = {option_value}")
            return

    role_radios = question.locator('[role="radio"]')
    role_count = role_radios.count()

    for i in range(role_count):
        item = role_radios.nth(i)
        aria_label = normalize_text(item.get_attribute("aria-label"))
        value_attr = normalize_text(item.get_attribute("value"))
        try:
            inner_txt = normalize_text(item.inner_text(timeout=1000))
        except Exception:
            inner_txt = ""

        candidates = [aria_label, value_attr, inner_txt]

        if option_value in candidates:
            item.click(force=True, timeout=10000)
            page.wait_for_timeout(400)
            log(f"Q{question_number}: rating/radio selected by role=radio = {option_value}")
            return

    text_items = question.locator('div, span, label, button')

    for i in range(text_items.count()):
        item = text_items.nth(i)
        try:
            if not item.is_visible():
                continue
        except Exception:
            pass

        try:
            txt = normalize_text(item.inner_text(timeout=500))
        except Exception:
            txt = ""

        aria_label = normalize_text(item.get_attribute("aria-label"))

        if txt == option_value or aria_label == option_value:
            try:
                item.click(force=True, timeout=10000)
                page.wait_for_timeout(400)
                log(f"Q{question_number}: selected by visible text fallback = {option_value}")
                return
            except Exception:
                pass

    for i in range(radio_count):
        radio = radios.nth(i)
        val = normalize_text(radio.get_attribute("value"))

        if option_value in val or val in option_value:
            radio.check(force=True, timeout=10000)
            page.wait_for_timeout(400)
            log(f"Q{question_number}: radio selected by partial input value = {val}")
            return

    available = []

    for i in range(radio_count):
        available.append(normalize_text(radios.nth(i).get_attribute("value")))

    for i in range(role_count):
        item = role_radios.nth(i)
        try:
            item_text = normalize_text(item.inner_text(timeout=500))
        except Exception:
            item_text = ""
        available.append("role-radio | aria-label=" + normalize_text(item.get_attribute("aria-label")) + " | text=" + item_text)

    fail_shot = os.path.join(SCREENSHOT_DIR, f"q{question_number}_radio_option_not_found_{CURRENT_CODE}_{RUN_TIME}.png")
    try:
        page.screenshot(path=fail_shot, full_page=True)
    except Exception:
        pass

    try:
        q_text = question.inner_text(timeout=3000)
    except Exception:
        q_text = ""

    log(f"Q{question_number}: Radio option not found. Screenshot: {fail_shot}")
    log(f"Q{question_number}: Required option: {option_value}")
    log(f"Q{question_number}: Available options: {available}")
    log(f"Q{question_number}: Question text preview: {q_text[:1500]}")
    raise Exception(f"Q{question_number}: Option not found: {option_value}. Available: {available}")


def upload_file_to_question(page, question_number, file_path):
    file_path = str(file_path or "").strip()

    if not file_path:
        log(f"Q{question_number}: file path is blank -> skipped")
        return

    if not os.path.exists(file_path):
        try:
            ensure_file_available(file_path)
        except Exception:
            log(f"Q{question_number}: WARNING file not found -> skipped: {file_path}")
            return

    question = scroll_to_question(page, question_number)
    file_locator = question.locator('input[type="file"]')

    if file_locator.count() == 0:
        fail_shot = os.path.join(SCREENSHOT_DIR, f"q{question_number}_file_input_not_found_{CURRENT_CODE}_{RUN_TIME}.png")
        try:
            page.screenshot(path=fail_shot, full_page=True)
        except Exception:
            pass
        log(f"Q{question_number}: File input not found. Screenshot: {fail_shot}")
        raise Exception(f"Q{question_number}: File input not found")

    file_input = file_locator.first
    file_input.set_input_files(file_path, timeout=30000)
    page.wait_for_timeout(8000)
    log(f"Q{question_number}: file uploaded = {file_path}")


def apply_question(page, ws, question_number, cfg):
    qtype = cfg["type"]

    if qtype in ["textarea", "date", "radio"]:
        value = get_value_from_mapping(ws, cfg)
    else:
        value = cfg.get("file", "")

    log(f"Processing Q{question_number} | type={qtype} | value/source preview={str(value)[:120]}")

    if qtype == "textarea":
        fill_textarea_question(page, question_number, value)
    elif qtype == "date":
        fill_date_question(page, question_number, value)
    elif qtype == "radio":
        select_radio_in_question(page, question_number, value)
    elif qtype == "file":
        upload_file_to_question(page, question_number, value)
    else:
        raise Exception(f"Unsupported type for Q{question_number}: {qtype}")

# ============================================================
# EMAIL RECEIPT + SUBMIT
# ============================================================

def check_email_receipt_box(page):
    if not SEND_EMAIL_RECEIPT:
        log("Email receipt checkbox disabled in config -> skipped")
        return

    log("Trying to check email receipt checkbox...")

    checkbox = page.locator('input[type="checkbox"]').first

    if checkbox.count() > 0:
        try:
            checkbox.scroll_into_view_if_needed(timeout=10000)
            try:
                is_checked = checkbox.is_checked()
            except Exception:
                is_checked = normalize_text(checkbox.get_attribute("aria-checked")).lower() == "true"

            if not is_checked:
                checkbox.check(force=True, timeout=10000)

            page.wait_for_timeout(500)
            log("Email receipt checkbox checked by input[type=checkbox]")
            return
        except Exception:
            pass

    receipt_text = page.locator('text="Send me an email receipt of my responses"')

    if receipt_text.count() > 0:
        try:
            receipt_text.first.scroll_into_view_if_needed(timeout=10000)
            receipt_text.first.click(force=True, timeout=10000)
            page.wait_for_timeout(500)
            log("Email receipt checkbox checked by visible text")
            return
        except Exception:
            pass

    receipt_any = page.locator('div, span, label').filter(has_text="Send me an email receipt of my responses")

    if receipt_any.count() > 0:
        try:
            receipt_any.first.scroll_into_view_if_needed(timeout=10000)
            receipt_any.first.click(force=True, timeout=10000)
            page.wait_for_timeout(500)
            log("Email receipt checkbox checked by text container")
            return
        except Exception:
            pass

    fail_shot = os.path.join(SCREENSHOT_DIR, f"email_receipt_checkbox_not_found_{CURRENT_CODE}_{RUN_TIME}.png")
    try:
        page.screenshot(path=fail_shot, full_page=True)
    except Exception:
        pass
    log(f"WARNING: Email receipt checkbox not found. Screenshot: {fail_shot}")


def submit_form(page):
    if not SUBMIT:
        log("SUBMIT = False -> Submit skipped")
        return

    log("SUBMIT = True -> Trying to submit form...")

    submit_locators = [
        page.get_by_role("button", name=re.compile(r"Submit|إرسال|ارسال|تقديم", re.I)),
        page.locator('button:has-text("Submit")'),
        page.locator('text="Submit"'),
    ]

    for btn in submit_locators:
        try:
            if btn.count() > 0 and btn.first.is_visible():
                btn.first.scroll_into_view_if_needed(timeout=10000)
                page.wait_for_timeout(500)
                btn.first.click(timeout=15000)
                page.wait_for_timeout(6000)
                log("Submit button clicked successfully")
                return
        except Exception:
            pass

    fail_shot = os.path.join(SCREENSHOT_DIR, f"submit_button_not_found_{CURRENT_CODE}_{RUN_TIME}.png")
    try:
        page.screenshot(path=fail_shot, full_page=True)
    except Exception:
        pass
    log(f"ERROR: Submit button not found. Screenshot: {fail_shot}")
    raise Exception("Submit button not found")

# ============================================================
# MAIN FORM RUNNER
# ============================================================

def run_visit_form(code):
    profile = VISIT_PROFILES[code]
    sheet_name = profile["sheet_name"]
    form_map = build_form_map(profile)

    log("========== START VISIT FORM FILL ==========")
    log(f"Visit code = {code}")
    log(f"Sheet name = {sheet_name}")
    log(f"Distributor = {profile['distributor']}")
    log(f"SUBMIT = {SUBMIT}")
    log(f"SEND_EMAIL_RECEIPT = {SEND_EMAIL_RECEIPT}")
    log(f"CLEAR_FORM_BEFORE_FILL = {CLEAR_FORM_BEFORE_FILL}")
    log(f"STRICT_CLEAR_FORM_REQUIRED = {STRICT_CLEAR_FORM_REQUIRED}")
    log(f"Excel file = {EXCEL_FILE}")

    ensure_file_available(EXCEL_FILE)

    wb = load_workbook(EXCEL_FILE, data_only=True, keep_vba=True)

    if sheet_name not in wb.sheetnames:
        raise Exception(f"Sheet not found: {sheet_name}. Available: {wb.sheetnames}")

    ws = wb[sheet_name]

    with sync_playwright() as p:
        context = p.chromium.launch_persistent_context(
            user_data_dir=USER_DATA_DIR,
            channel="msedge" if not CLOUD_MODE else "chromium",
            headless=CLOUD_MODE,
            viewport={"width": 1366, "height": 768},
            args=[
                "--disable-blink-features=AutomationControlled",
                "--start-maximized",
            ],
        )

        page = context.new_page()

        log("Opening form...")
        page.goto(FORM_URL, wait_until="domcontentloaded", timeout=60000)

        total_questions = wait_until_questions_ready(page)
        log(f"Total questions found on page: {total_questions}")

        clear_form_before_fill(page)

        for question_number in sorted(form_map.keys()):
            apply_question(page, ws, question_number, form_map[question_number])

        check_email_receipt_box(page)

        log("Taking final screenshot before submit...")
        page.screenshot(path=SCREENSHOT_FILE, full_page=True)
        log(f"Screenshot saved: {SCREENSHOT_FILE}")

        submit_form(page)

        if SUBMIT:
            log("Form submitted because SUBMIT = True")
        else:
            log("No submit clicked. This is test mode only.")

        print("")
        print(f"DONE - Form filled from Excel for code: {code}")
        print("")
        print(f"SUBMIT = {SUBMIT}")
        print("No submit was clicked." if not SUBMIT else "Form was submitted.")
        print("")
        print("Screenshot:")
        print(SCREENSHOT_FILE)
        print("")
        print("Log:")
        print(LOG_FILE)
        print("")
        print("لو التست مظبوط وعايز تبعت فعليًا:")
        print("غير السطر:")
        print("SUBMIT = False")
        print("إلى:")
        print("SUBMIT = True")
        print("")

        if not CLOUD_MODE:
            input("Press Enter to close browser...")
        context.close()

    log("========== END VISIT FORM FILL ==========")

# ============================================================
# MAIN
# ============================================================

def main():
    log("========== START MASTER VISIT AUTOMATION ==========")

    try:
        code = refresh_workbook_and_get_today_code()
        set_run_files(code)

        log(f"Running full integrated script for code: {code}")
        run_visit_form(code)

        log("========== MASTER VISIT AUTOMATION COMPLETED SUCCESSFULLY ==========")

    except Exception as e:
        log(f"ERROR: {e}")
        log("========== MASTER VISIT AUTOMATION FAILED ==========")
        if not CLOUD_MODE:
            input("Press Enter to close...")
        raise


if __name__ == "__main__":
    main()
